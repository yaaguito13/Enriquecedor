"""
scraper.py — Agente de scraping empresarial mejorado
Filtros: solo empresas reales españolas, sin redes sociales ni plataformas
"""

import re
import time
import random
import logging
import threading
from html import unescape
from urllib.parse import urljoin, urlparse, quote_plus, unquote
from dataclasses import dataclass, asdict
from typing import Optional
import requests
from bs4 import BeautifulSoup

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger(__name__)

# ── Rotación de User-Agents reales (Chrome / Firefox / Edge modernos) ──
_USER_AGENTS = [
    # Chrome 124 Windows
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    # Chrome 123 macOS
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
    # Firefox 125 Windows
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:125.0) Gecko/20100101 Firefox/125.0",
    # Firefox 124 Linux
    "Mozilla/5.0 (X11; Linux x86_64; rv:124.0) Gecko/20100101 Firefox/124.0",
    # Edge 124 Windows
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36 Edg/124.0.0.0",
    # Safari 17 macOS
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 14_4_1) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.4.1 Safari/605.1.15",
]

def _make_headers(referer: str = "") -> dict:
    """
    Genera headers que simulan un navegador real con rotación de UA.
    Incluye sec-fetch-* y Accept-Encoding que los bots suelen omitir.
    """
    ua = random.choice(_USER_AGENTS)
    is_firefox = "Firefox" in ua

    headers = {
        "User-Agent": ua,
        "Accept": (
            "text/html,application/xhtml+xml,application/xml;q=0.9,"
            "image/avif,image/webp,image/apng,*/*;q=0.8"
            if not is_firefox else
            "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8"
        ),
        "Accept-Language": "es-ES,es;q=0.9,en;q=0.7",
        "Accept-Encoding": "gzip, deflate, br",
        "Connection": "keep-alive",
        "Upgrade-Insecure-Requests": "1",
        "Sec-Fetch-Dest": "document",
        "Sec-Fetch-Mode": "navigate",
        "Sec-Fetch-Site": "none" if not referer else "same-origin",
        "Sec-Fetch-User": "?1",
        "DNT": "1",
    }
    if referer:
        headers["Referer"] = referer
    if not is_firefox:
        headers["Sec-Ch-Ua"] = '"Chromium";v="124", "Google Chrome";v="124", "Not-A.Brand";v="99"'
        headers["Sec-Ch-Ua-Mobile"] = "?0"
        headers["Sec-Ch-Ua-Platform"] = '"Windows"'
    return headers

# Headers base para la sesión (se sobreescriben por petición)
HEADERS = _make_headers()

REQUEST_TIMEOUT = 12
DELAY_BETWEEN_REQUESTS = (1.5, 3.5)  # Más conservador para evitar bloqueos

# ──────────────────────────────────────────────────────────────────────
# REGEX DE EMAILS — captura emails en texto/HTML y valida después:
#   · No captura paths de ficheros (algo@algo.png)
#   · No captura versiones (1.0@...)
#   · Requiere TLD de 2-24 chars y que el dominio tenga al menos un punto
#   · Permite emails dentro de atributos y JS inline
# ──────────────────────────────────────────────────────────────────────
EMAIL_REGEX = re.compile(
    r"""
    (?<![A-Z0-9._%+\-])
    [A-Z0-9]
    [A-Z0-9._%+\-]{0,62}
    [A-Z0-9]
    @
    (?:[A-Z0-9](?:[A-Z0-9\-]{0,61}[A-Z0-9])?\.)+
    [A-Z]{2,24}
    (?![A-Z0-9._%+\-])
    """,
    re.VERBOSE | re.IGNORECASE,
)

# ──────────────────────────────────────────────────────────────────────
# REGEX DE TELÉFONOS ESPAÑOLES — cubre todos los formatos:
#   Estrategia: capturar el número completo de 9 dígitos con separadores
#   opcionales en cualquier posición, validando que empiece por 6-9.
#   Se usa _clean_phone() después para normalizar y validar longitud.
#
#   Formatos cubiertos:
#     +34 912 345 678  · +34912345678  · 0034 91 234 56 78
#     91 234 56 78     · 93-234-5678   · 634.567.890
#     634 567 890      · 6345 678 90   · (93) 234 5678
# ──────────────────────────────────────────────────────────────────────
PHONE_REGEX = re.compile(
    r"""
    (?<!\d)
    (?:\+34|0034|34)?
    [\s().\-]*
    (?:[6789]\d|900|901|902)
    (?:[\s().\-]*\d){7}
    (?!\d)
    """,
    re.VERBOSE,
)

OBFUSCATED_EMAIL_REGEX = re.compile(
    r"""
    (?<!\w)
    ([A-Z0-9._%+\-]{1,64})
    \s*(?:\[?\s*(?:at|arroba)\s*\]?|\(at\)|\{at\})\s*
    ([A-Z0-9.\-]+\s*(?:\[?\s*(?:dot|punto)\s*\]?|\(dot\)|\{dot\}|\.)\s*[A-Z]{2,24})
    """,
    re.VERBOSE | re.IGNORECASE,
)

BLOCKED_DOMAINS = {
    "linkedin.com","es.linkedin.com","il.linkedin.com","twitter.com","x.com","t.co",
    "facebook.com","fb.com","instagram.com","youtube.com","youtu.be",
    "tiktok.com","pinterest.com","pinterest.es","flickr.com","vimeo.com",
    "apps.apple.com","play.google.com","microsoft.com",
    "google.com","google.es","googleapis.com","policies.google.com",
    "accounts.google.com","maps.google.com","goo.gl",
    "europages.com","europages.es","europages.co.uk","visable.com",
    "kompass.com","einforma.com","axesor.es","infoempresa.com",
    "empresite.eleconomista.es","ranking-empresas.eleconomista.es","eleconomista.es",
    "issuu.com","slideshare.net","scribd.com","medium.com","substack.com",
    "wordpress.com","blogspot.com","blogger.com",
    "whatsapp.com","telegram.org","amazon.com","amazon.es","ebay.es","ebay.com",
    "wikipedia.org","wikimedia.org","github.com","gitlab.com",
    "zoom.us","dropbox.com","drive.google.com","onedrive.live.com",
    "conwize.io","entramos.com","mis.com.es",
    "itunes.apple.com","boe.es","ine.es","seg-social.es","agenciatributaria.es",
    "cloudflare.com","jsdelivr.net","fonts.googleapis.com",
    "seopan.es",  # Asociación, no directorio de empresas
}

BLOCKED_NAME_KEYWORDS = {
    "linkedin","twitter","facebook","instagram","youtube","tiktok",
    "flickr","issuu","vimeo","pinterest","google","apple","microsoft","amazon",
    "antes de ir a youtube","inicio de sesión","log in","sign in","app store", "app\\u00a0store",
    "google play","play store","europages","visable","kompass",
    "política de privacidad","términos","aviso legal","cookies",
    "help center","centro de ayuda","soporte","bme",
    "dominios disponibles","entramos","inicio de sesión en linkedin",
}

# ──────────────────────────────────────────────────────────────────────
# Helpers globales — Cloudflare, JSON-LD, Playwright
# ──────────────────────────────────────────────────────────────────────

def _decode_cloudflare_email(encoded: str) -> Optional[str]:
    """
    Decodifica un email ofuscado por Cloudflare Email Address Obfuscation.
    El formato es una cadena hexadecimal donde:
    - Los dos primeros bytes son la clave XOR
    - El resto son los bytes del email codificados con XOR contra esa clave

    Ejemplo: data-cfemail="a1cde4e0..." → "info@empresa.es"
    """
    try:
        if not encoded or len(encoded) < 4:
            return None
        # Limpiar espacios y convertir a bytes
        encoded = encoded.strip().replace(" ", "")
        if len(encoded) % 2 != 0:
            return None
        data = bytes.fromhex(encoded)
        key = data[0]
        decoded = "".join(chr(b ^ key) for b in data[1:])
        return decoded if "@" in decoded else None
    except Exception:
        return None


def _extract_emails_from_obj(obj, _depth: int = 0) -> list[str]:
    """
    Recorre recursivamente un objeto JSON (dict/list) y extrae
    cualquier string que parezca un email. Límite de profundidad: 6.
    Útil para JSON-LD / Schema.org.
    """
    if _depth > 6:
        return []
    emails = []
    if isinstance(obj, str):
        if "@" in obj and "." in obj:
            emails.append(obj.strip())
    elif isinstance(obj, dict):
        # Campos con alta probabilidad de contener email
        priority_keys = {"email", "contactEmail", "contactPoint", "Email"}
        for k, v in obj.items():
            if k in priority_keys and isinstance(v, str) and "@" in v:
                emails.append(v.strip())
            else:
                emails.extend(_extract_emails_from_obj(v, _depth + 1))
    elif isinstance(obj, list):
        for item in obj:
            emails.extend(_extract_emails_from_obj(item, _depth + 1))
    return emails


def _extract_phones_from_obj(obj, _depth: int = 0) -> list[str]:
    """
    Recorre recursivamente un objeto JSON y extrae teléfonos.
    Útil para JSON-LD / Schema.org (campo 'telephone').
    """
    if _depth > 6:
        return []
    phones = []
    if isinstance(obj, str):
        if re.search(r"[\d\+]", obj) and len(obj) <= 20:
            phones.append(obj.strip())
    elif isinstance(obj, dict):
        priority_keys = {"telephone", "phone", "tel", "faxNumber", "Telephone"}
        for k, v in obj.items():
            if k in priority_keys and isinstance(v, str):
                phones.append(v.strip())
            else:
                phones.extend(_extract_phones_from_obj(v, _depth + 1))
    elif isinstance(obj, list):
        for item in obj:
            phones.extend(_extract_phones_from_obj(item, _depth + 1))
    return phones


def _unique(items: list[str]) -> list[str]:
    """Devuelve la lista sin duplicados conservando el orden."""
    seen = set()
    result = []
    for item in items:
        if item and item not in seen:
            seen.add(item)
            result.append(item)
    return result


def _decode_text_for_contact_search(text: str) -> str:
    """Normaliza texto/HTML antes de buscar contactos."""
    text = unescape(text or "")
    text = (
        text
        .replace("\\u0040", "@")
        .replace("\\x40", "@")
        .replace("%40", "@")
        .replace("&#64;", "@")
    )
    text = re.sub(r"\s*(?:\[\s*at\s*\]|\(\s*at\s*\)|\{\s*at\s*\}|(?<!\w)arroba(?!\w))\s*", "@", text, flags=re.I)
    text = re.sub(r"\s*(?:\[\s*dot\s*\]|\(\s*dot\s*\)|\{\s*dot\s*\}|(?<!\w)punto(?!\w))\s*", ".", text, flags=re.I)
    return text


def _extract_emails_from_text(text: str) -> list[str]:
    """Extrae emails normales y ofuscados de un bloque de texto o HTML."""
    text = _decode_text_for_contact_search(text)
    candidates = list(EMAIL_REGEX.findall(text))
    for local, domain in OBFUSCATED_EMAIL_REGEX.findall(text):
        domain = re.sub(r"\s*(?:\[?\s*(?:dot|punto)\s*\]?|\(dot\)|\{dot\})\s*", ".", domain, flags=re.I)
        domain = re.sub(r"\s+", "", domain)
        candidates.append(f"{local}@{domain}")
    return _unique([c for e in candidates if (c := _clean_email(e))])


def _extract_phones_from_text(text: str) -> list[str]:
    """Extrae teléfonos españoles con y sin separadores de un bloque de texto."""
    text = _decode_text_for_contact_search(text)
    candidates = PHONE_REGEX.findall(text)
    return _unique([c for p in candidates if (c := _clean_phone(p))])


def _normalize_url(url: str) -> str:
    """Acepta dominios escritos como empresa.es y los convierte a URL navegable."""
    url = (url or "").strip()
    if not url:
        return ""
    if url.startswith("//"):
        return "https:" + url
    if not urlparse(url).scheme:
        return "https://" + url
    return url


# ──────────────────────────────────────────────────────────────────────
# FALLBACK: Playwright (headless browser)
# ──────────────────────────────────────────────────────────────────────
#
# Usa este módulo cuando requests devuelve HTML vacío o bloqueado por JS.
# Playwright renderiza la página completa como un navegador real.
#
# INSTALACIÓN:
#   pip install playwright
#   playwright install chromium
#
# INTEGRACIÓN en ScraperAgent._get():
#
#   def _get_with_playwright(self, url: str) -> Optional[BeautifulSoup]:
#       """Fallback con Playwright para páginas que requieren JavaScript."""
#       try:
#           from playwright.sync_api import sync_playwright
#           with sync_playwright() as p:
#               browser = p.chromium.launch(headless=True)
#               context = browser.new_context(
#                   user_agent=random.choice(_USER_AGENTS),
#                   locale="es-ES",
#                   extra_http_headers={"Accept-Language": "es-ES,es;q=0.9"},
#               )
#               page = context.new_page()
#               # Bloquear imágenes y fuentes para ir más rápido
#               page.route("**/*.{png,jpg,jpeg,gif,svg,ico,woff,woff2}",
#                          lambda route: route.abort())
#               page.goto(url, wait_until="networkidle", timeout=20_000)
#               # Esperar a que aparezca algún texto en el body
#               page.wait_for_selector("body", timeout=10_000)
#               html = page.content()
#               browser.close()
#               return BeautifulSoup(html, "html.parser")
#       except Exception as e:
#           log.warning(f"Playwright falló en {url}: {e}")
#           return None
#
# CUÁNDO ACTIVARLO:
#   Detectar respuesta bloqueada en _get() y llamar a _get_with_playwright():
#
#   soup = self._get(url)
#   if soup is None or len(soup.get_text(strip=True)) < 300:
#       soup = self._get_with_playwright(url)
#
# NOTA: Playwright es ~10x más lento que requests. Úsalo solo como
# fallback después de que requests falle, no como método principal.
# ──────────────────────────────────────────────────────────────────────

SECTOR_QUERIES = {
    "Tecnología": [
        'directorio empresas tecnología software España',
        'listado startups tecnológicas España contacto web',
        'empresas desarrollo software España catálogo',
        'agencias desarrollo web España email teléfono',
        'proveedores IT tecnología España pymes',
    ],
    "Marketing": [
        'directorio agencias marketing digital España',
        'listado agencias publicidad España contacto',
        'agencias marketing online España catálogo web',
        'empresas branding comunicación España pymes',
        'agencias SEO SEM España web contacto',
    ],
    "Diseño": [
        'directorio estudios diseño gráfico España',
        'listado agencias diseño web España contacto',
        'estudios diseño industrial España catálogo',
        'empresas diseño UX UI España pymes web',
        'estudios creativos diseño España teléfono email',
    ],
    "Construcción": [
        'directorio constructoras España web contacto',
        'listado empresas construcción reformas España email',
        'constructoras obra civil España catálogo empresas',
        'empresas reformas construcción España pymes',
        'promotoras constructoras España web teléfono',
    ],
    "Salud": [
        'directorio clínicas privadas España contacto web',
        'listado centros médicos privados España email',
        'clínicas hospitales privados España teléfono web',
        'empresas servicios médicos España catálogo pymes',
        'centros salud privados España directorio',
    ],
    "Legal": [
        'directorio despachos abogados España web',
        'listado bufetes jurídicos España contacto email',
        'asesorías jurídicas España catálogo empresas web',
        'despachos abogados España teléfono contacto',
        'firmas legales abogados España pymes directorio',
    ],
    "Educación": [
        'directorio academias formación España web',
        'listado centros educativos privados España contacto',
        'academias formación profesional España email teléfono',
        'empresas elearning formación España catálogo',
        'colegios privados academias España directorio',
    ],
    "Restauración": [
        'directorio restaurantes España web contacto',
        'listado cadenas restauración España web oficial',
        'empresas hostelería restauración España catálogo',
        'franquicias restauración España directorio email',
        'grupos restauración hostelería España teléfono',
    ],
    "Inmobiliaria": [
        'directorio agencias inmobiliarias España web',
        'listado inmobiliarias España contacto email',
        'promotoras inmobiliarias España catálogo empresas',
        'agencias propiedades España directorio pymes',
        'gestoras inmobiliarias España web teléfono',
    ],
    "Industria": [
        'directorio fabricantes industriales España web',
        'listado empresas manufactura industria España contacto',
        'fabricantes sector industrial España catálogo',
        'proveedores industriales España directorio',
        'empresas industria España email teléfono web',
    ],
    "Otros": [
        'directorio empresas servicios España web',
        'listado pymes España directorio contacto email',
        'catálogo empresas servicios profesionales España',
        'empresas servicios España web teléfono',
    ],
}

SECTOR_KEYWORDS = {
    "Tecnología": ["tecnología", "software", "informática", "desarrollo", "web", "app", "it ", "digital", "tecnológico", "programación", "sistemas"],
    "Marketing": ["marketing", "publicidad", "seo", "sem", "agencia", "comunicación", "digital", "branding", "diseño"],
    "Diseño": ["diseño", "gráfico", "estudio", "creativo", "interiorismo", "arquitectura", "web", "ux", "ui"],
    "Construcción": ["construcción", "reformas", "obras", "edificación", "arquitectura", "constructora", "promotora", "inmobiliaria", "instalaciones"],
    "Salud": ["salud", "clínica", "médico", "hospital", "dental", "fisioterapia", "psicología", "medicina", "pacientes", "tratamiento", "estética"],
    "Legal": ["legal", "abogados", "asesoría", "jurídico", "bufete", "derecho", "consultoría", "gestoría", "leyes"],
    "Educación": ["educación", "formación", "academia", "colegio", "escuela", "cursos", "universidad", "profesores", "alumnos", "clases"],
    "Restauración": ["restaurante", "hostelería", "bar", "cafetería", "comida", "gastronomía", "catering", "hotel", "alimentación", "menú"],
    "Inmobiliaria": ["inmobiliaria", "propiedades", "pisos", "casas", "venta", "alquiler", "promotora", "inmuebles", "agencia"],
    "Industria": ["industria", "fábrica", "fabricación", "industrial", "maquinaria", "producción", "ingeniería", "suministros", "materiales"],
}

@dataclass
class Company:
    empresa: str
    sector: str
    web: Optional[str] = None
    email: Optional[str] = None
    telefono: Optional[str] = None
    fuente: Optional[str] = None

    def key(self) -> str:
        if self.web:
            try:
                return urlparse(self.web).netloc.lower().replace("www.", "")
            except Exception:
                pass
        return self.empresa.lower().strip()[:40]


def _is_blocked_domain(url: str) -> bool:
    try:
        url = _normalize_url(url)
        netloc = urlparse(url).netloc.lower()
        for blocked in BLOCKED_DOMAINS:
            if netloc == blocked or netloc.endswith("." + blocked):
                return True
    except Exception:
        pass
    return False


def _is_blocked_name(name: str) -> bool:
    import unicodedata
    lower = unicodedata.normalize("NFKC", name).lower()
    return any(kw in lower for kw in BLOCKED_NAME_KEYWORDS)


def _looks_like_company_page(soup: BeautifulSoup, url: str) -> bool:
    bad_patterns = [
        r"/login", r"/logout", r"/signup", r"/register",
        r"/privacy", r"/terms", r"/cookies", r"/legal",
        r"\?logout=", r"shareArticle", r"/share\?",
    ]
    path_q = urlparse(url).path + "?" + (urlparse(url).query or "")
    for pat in bad_patterns:
        if re.search(pat, path_q, re.I):
            return False
    text = soup.get_text(strip=True)
    return len(text) >= 200


def _clean_phone(phone: str) -> Optional[str]:
    """Normaliza y valida un teléfono español. Devuelve +34XXXXXXXXX o None."""
    # Eliminar todo salvo dígitos y el signo +
    digits = re.sub(r"[^\d+]", "", phone)

    # Normalizar prefijos
    if digits.startswith("+340"):          # caso raro +340XX
        return None
    if digits.startswith("0034"):
        digits = "+34" + digits[4:]
    elif digits.startswith("+34"):
        pass                               # ya correcto
    elif digits.startswith("34") and len(digits) == 11:
        digits = "+34" + digits[2:]
    elif not digits.startswith("+") and len(digits) == 9:
        # Verificar que empieza por 6, 7, 8 o 9
        if digits[0] not in "6789":
            return None
        digits = "+34" + digits
    else:
        return None

    # Validar longitud final: +34 + 9 dígitos = 12 chars
    if len(digits) != 12:
        return None

    # Descartar números de prueba obvios (666666666, 000000000, etc.)
    body = digits[3:]
    if len(set(body)) <= 2:               # demasiados dígitos iguales
        return None

    return digits


def _clean_email(email: str) -> Optional[str]:
    """Valida y normaliza un email. Devuelve minúsculas o None."""
    email = email.strip().lower()

    # Extensiones que son ficheros, no emails
    bad_ext = {
        "png","jpg","jpeg","gif","svg","webp","ico","bmp","tiff",
        "woff","woff2","ttf","eot","otf",
        "css","js","jsx","ts","tsx","json","xml","yaml","yml",
        "pdf","doc","docx","xls","xlsx","zip","rar","gz","tar",
        "mp4","mp3","avi","mov","wav",
        "map","min","bundle","chunk",
    }
    tld = email.rsplit(".", 1)[-1] if "." in email else ""
    if tld in bad_ext:
        return None

    # Longitud razonable
    if len(email) < 6 or len(email) > 100:
        return None

    # Debe tener exactamente un @
    parts = email.split("@")
    if len(parts) != 2:
        return None
    local, domain = parts

    # Parte local: al menos 1 char
    if len(local) < 1:
        return None

    # Dominio: al menos un punto y TLD de al menos 2 chars
    domain_parts = domain.split(".")
    if len(domain_parts) < 2 or len(domain_parts[-1]) < 2:
        return None
    if any(not p or p.startswith("-") or p.endswith("-") for p in domain_parts):
        return None

    # Dominios de plataformas falsas o de prueba
    bad_domains = {
        "sentry.io","example.com","example.es","test.com","test.es",
        "domain.com","yoursite.com","tuempresa.com","empresa.com",
        "correo.com","email.com","mail.com","noreply.com",
        "visable.com","europages.com","kompass.com",
        "wixpress.com","squarespace.com",  # internos de plataformas web
    }
    if domain in bad_domains:
        return None

    # Partes locales que indican no-reply o prueba
    bad_local_prefixes = {
        "noreply","no-reply","no_reply","donotreply","do-not-reply",
        "mailer-daemon","postmaster","bounce","unsubscribe",
    }
    if local in bad_local_prefixes:
        return None

    # Evitar emails que son en realidad números de versión
    if re.match(r"^\d+\.\d+", local):
        return None
    if ".." in local or ".." in domain:
        return None
    if local.startswith(".") or local.endswith("."):
        return None

    return email


class ScraperAgent:
    def __init__(self, sector: str, keywords: str = "", status_cb=None):
        self.sector = sector
        self.keywords = keywords.strip()
        self.status_cb = status_cb or (lambda msg, **kw: None)
        self.session = requests.Session()
        self.session.headers.update(HEADERS)
        self.companies: dict[str, Company] = {}
        self.sources_visited: set[str] = set()
        self.errors: list[str] = []
        self._stop_event = threading.Event()

    def stop(self):
        self._stop_event.set()

    def _stopped(self) -> bool:
        return self._stop_event.is_set()

    def _emit(self, message: str, **kwargs):
        self.status_cb(message, **kwargs)

    @property
    def stats(self) -> dict:
        c = list(self.companies.values())
        return {
            "fuentes": len(self.sources_visited),
            "empresas": len(c),
            "emails": sum(1 for x in c if x.email),
            "telefonos": sum(1 for x in c if x.telefono),
        }

    def _get(self, url: str, referer: str = "") -> Optional[BeautifulSoup]:
        """
        GET con:
        - Rotación de User-Agent por petición
        - Retry automático ante 429 / 503 con backoff exponencial
        - Detección de respuesta vacía o bloqueada
        """
        if self._stopped() or _is_blocked_domain(url):
            return None

        max_retries = 3
        for attempt in range(max_retries):
            try:
                # Headers frescos con UA rotado en cada intento
                self.session.headers.update(_make_headers(referer=referer or url))
                time.sleep(random.uniform(*DELAY_BETWEEN_REQUESTS))

                resp = self.session.get(
                    url,
                    timeout=REQUEST_TIMEOUT,
                    allow_redirects=True,
                )

                # Detectar bloqueos por rate-limit → esperar y reintentar
                if resp.status_code == 429:
                    wait = int(resp.headers.get("Retry-After", 10)) + random.uniform(2, 5)
                    log.warning(f"429 en {url} — esperando {wait:.0f}s")
                    time.sleep(wait)
                    continue

                # 403 en primer intento: reintentar con UA diferente
                if resp.status_code == 403 and attempt < max_retries - 1:
                    log.debug(f"403 en {url}, reintentando con UA diferente")
                    time.sleep(random.uniform(3, 6))
                    continue

                # Otros errores HTTP
                resp.raise_for_status()

                # Detectar redireccionado a dominio bloqueado
                if _is_blocked_domain(resp.url):
                    return None

                # Detectar respuesta vacía o page bloqueada por JS
                content = resp.text
                if len(content.strip()) < 100:
                    log.debug(f"Respuesta vacía/bloqueada en {url}")
                    return None

                # Señales de que la página requiere JS para mostrar contenido
                js_only_signals = [
                    "enable javascript",
                    "you need to enable javascript",
                    "please enable javascript",
                    "browser does not support javascript",
                    "this site requires javascript",
                ]
                content_lower = content[:2000].lower()
                if any(s in content_lower for s in js_only_signals):
                    log.debug(f"Página requiere JS: {url}")
                    # No devolvemos None — puede haber algo de contenido útil
                    # pero sí anotamos para el fallback con Playwright

                return BeautifulSoup(content, "html.parser")

            except requests.exceptions.Timeout:
                log.debug(f"Timeout en {url} (intento {attempt+1})")
                time.sleep(2 ** attempt)
            except requests.exceptions.TooManyRedirects:
                log.debug(f"Demasiados redirects en {url}")
                return None
            except requests.exceptions.ConnectionError:
                log.debug(f"Error de conexión en {url}")
                time.sleep(2 ** attempt)
            except requests.exceptions.HTTPError as e:
                log.debug(f"HTTP {e.response.status_code} en {url}")
                return None
            except Exception as e:
                self.errors.append(f"{url}: {e}")
                return None

        return None  # Agotados los reintentos

    def _search_urls(self, query: str) -> list[str]:
        from ddgs import DDGS
        full_query = f"{query} España"
        if self.keywords:
            full_query += f" {self.keywords}"
        urls = []
        try:
            with DDGS() as ddgs:
                results = ddgs.text(full_query, region="es-es", max_results=10)
                for r in results:
                    href = r.get("href", "")
                    if href.startswith("http") and not _is_blocked_domain(href):
                        urls.append(href)
        except Exception as e:
            log.debug(f"DDGS error en _search_urls: {e}")
        return list(dict.fromkeys(urls))

    # ── Extracción de emails ──────────────────────────────────────────
    def _extract_emails(self, text: str) -> list[str]:
        return _extract_emails_from_text(text)

    def _extract_emails_from_soup(self, soup: BeautifulSoup) -> list[str]:
        """
        Pipeline completo de extracción de emails desde el HTML parseado.
        Orden de prioridad: mailto: hrefs → JSON-LD → Cloudflare decode → regex en texto
        """
        found: list[str] = []

        # ── 1. mailto: en atributos href (la fuente más fiable) ──────
        for a in soup.find_all("a", href=True):
            href = a["href"]
            if href.lower().startswith("mailto:"):
                raw = href[7:].split("?")[0].strip()  # quitar ?subject= etc.
                c = _clean_email(raw)
                if c and c not in found:
                    found.append(c)

        # ── 2. JSON-LD / Schema.org ──────────────────────────────────
        # Muchas webs corporativas publican contactInfo en JSON-LD
        for script in soup.find_all("script", type="application/ld+json"):
            try:
                import json
                data = json.loads(script.string or "")
                # Buscar email en cualquier nivel del JSON
                emails_in_json = _extract_emails_from_obj(data)
                for e in emails_in_json:
                    c = _clean_email(e)
                    if c and c not in found:
                        found.append(c)
            except Exception:
                pass

        # ── 3. Cloudflare email obfuscation ──────────────────────────
        # Cloudflare reemplaza emails por <a data-cfemail="HEXSTRING">
        for el in soup.select("[data-cfemail]"):
            decoded = _decode_cloudflare_email(el.get("data-cfemail", ""))
            if decoded:
                c = _clean_email(decoded)
                if c and c not in found:
                    found.append(c)

        # Variante: __cf_email__ en spans dentro de <a>
        for el in soup.select(".__cf_email__"):
            encoded = el.get("data-cfemail", "")
            if encoded:
                decoded = _decode_cloudflare_email(encoded)
                if decoded:
                    c = _clean_email(decoded)
                    if c and c not in found:
                        found.append(c)

        # ── 4. Atributos data-email o data-mail ─────────────────────
        for el in soup.select("[data-email],[data-mail],[data-contact-email]"):
            for attr in ("data-email", "data-mail", "data-contact-email"):
                val = el.get(attr, "")
                if val:
                    c = _clean_email(val)
                    if c and c not in found:
                        found.append(c)

        # ── 5. Regex sobre el texto plano ───────────────────────────
        text = soup.get_text(separator=" ", strip=True)
        for e in self._extract_emails(text):
            if e not in found:
                found.append(e)

        # ── 6. Regex sobre el HTML crudo (captura emails en JS inline) ─
        # Ej: var email = "info@empresa.es"  o  'mailto:info@empresa.es'
        raw_html = str(soup)
        for e in _extract_emails_from_text(raw_html):
            if e not in found:
                found.append(e)

        return found

    # ── Extracción de teléfonos ───────────────────────────────────────
    def _extract_phones(self, text: str) -> list[str]:
        return _extract_phones_from_text(text)

    def _extract_phones_from_soup(self, soup: BeautifulSoup) -> list[str]:
        """
        Pipeline completo de extracción de teléfonos desde el HTML parseado.
        Orden: tel: hrefs → JSON-LD → atributos data-phone → regex en texto
        """
        found: list[str] = []

        # ── 1. tel: en atributos href ─────────────────────────────────
        for a in soup.find_all("a", href=True):
            href = a["href"]
            if href.lower().startswith("tel:"):
                raw = href[4:].strip()
                c = _clean_phone(raw)
                if c and c not in found:
                    found.append(c)

        # ── 2. JSON-LD / Schema.org ───────────────────────────────────
        for script in soup.find_all("script", type="application/ld+json"):
            try:
                import json
                data = json.loads(script.string or "")
                phones_in_json = _extract_phones_from_obj(data)
                for p in phones_in_json:
                    c = _clean_phone(p)
                    if c and c not in found:
                        found.append(c)
            except Exception:
                pass

        # ── 3. Atributos data-phone, data-tel, data-telefono ─────────
        for el in soup.select("[data-phone],[data-tel],[data-telefono],[data-telephone]"):
            for attr in ("data-phone","data-tel","data-telefono","data-telephone"):
                val = el.get(attr, "")
                if val:
                    c = _clean_phone(val)
                    if c and c not in found:
                        found.append(c)

        # ── 4. Regex sobre texto plano ────────────────────────────────
        text = soup.get_text(separator=" ", strip=True)
        for p in self._extract_phones(text):
            if p not in found:
                found.append(p)

        # ── 5. Regex sobre HTML crudo (teléfonos en atributos/JS) ─────
        for p in self._extract_phones(str(soup)):
            if p not in found:
                found.append(p)

        return found

    def _find_contact_url(self, base_url: str, soup: BeautifulSoup) -> Optional[str]:
        """
        Localiza la URL de la página de contacto usando un sistema de puntuación.
        Prioriza las coincidencias más específicas y descarta secciones de navegación
        genéricas.
        """
        # Palabras clave ordenadas por relevancia (más puntos = más relevante)
        KEYWORD_SCORES = {
            # Muy específico — casi seguro que es la página de contacto
            "contacto":         10,
            "contact":          10,
            "contactar":        10,
            "contáctanos":       9,
            "contáctenos":       9,
            "contactenos":       9,
            "get in touch":      9,
            # Específico — probable página de contacto o legal con datos
            "aviso legal":       7,
            "aviso-legal":       7,
            "legal notice":      7,
            "sobre nosotros":    6,
            "sobre-nosotros":    6,
            "quiénes somos":     6,
            "quienes-somos":     6,
            "about us":          6,
            "about-us":          6,
            "about":             4,
            "empresa":           4,
            "who we are":        4,
            # Menos específico — puede tener datos pero con menos probabilidad
            "información":       3,
            "info":              2,
            "team":              2,
            "equipo":            2,
        }

        candidates: list[tuple[int, str]] = []  # (score, url)

        for a in soup.find_all("a", href=True):
            href_raw = a["href"]
            href = href_raw.lower().strip()

            # Ignorar anchors, JS, mailto, tel
            if not href or href.startswith(("#", "javascript:", "mailto:", "tel:")):
                continue

            full_url = urljoin(base_url, href_raw)
            if _is_blocked_domain(full_url):
                continue

            # Evitar URLs que claramente no son de contacto
            skip_paths = [
                "/blog/", "/news/", "/noticias/", "/post/", "/tag/",
                "/category/", "/categoria/", "/producto/", "/product/",
                "/tienda/", "/shop/", "/carrito/", "/cart/",
                "/login", "/register", "/signup", "/logout",
                "/privacy", "/cookies", "/sitemap",
            ]
            if any(sp in full_url.lower() for sp in skip_paths):
                continue

            # Calcular score sumando keywords encontradas en texto y href
            link_text = a.get_text(strip=True).lower()
            score = 0
            for kw, pts in KEYWORD_SCORES.items():
                if kw in href or kw in link_text:
                    score += pts

            if score > 0:
                candidates.append((score, full_url))

        if not candidates:
            return None

        # Devolver la URL con mayor puntuación (ya vista o no)
        candidates.sort(key=lambda x: x[0], reverse=True)
        for _, url in candidates:
            if url != base_url:  # no volver a la misma página
                return url
        return None

    def _extract_name(self, soup: BeautifulSoup, url: str) -> Optional[str]:
        for prop in ["og:site_name", "og:title"]:
            m = soup.find("meta", property=prop)
            if m and m.get("content", "").strip():
                name = m["content"].strip()
                if 2 < len(name) < 80 and not _is_blocked_name(name):
                    return name
        title = soup.find("title")
        if title:
            name = title.get_text(strip=True)
            for sep in [" | "," - "," – "," :: "," · "," — "]:
                if sep in name:
                    name = name.split(sep)[0].strip()
            if 2 < len(name) < 80 and not _is_blocked_name(name):
                return name
        h1 = soup.find("h1")
        if h1:
            name = h1.get_text(strip=True)
            if 2 < len(name) < 80 and not _is_blocked_name(name):
                return name
        netloc = urlparse(url).netloc.replace("www.", "")
        dn = netloc.split(".")[0].title()
        return dn if dn and not _is_blocked_name(dn) else None

    def _process_company_url(self, url: str, source: str):
        if self._stopped() or url in self.sources_visited or _is_blocked_domain(url):
            return
        self.sources_visited.add(url)
        soup = self._get(url, referer=source)
        if not soup or not _looks_like_company_page(soup, url):
            return
            
        if self.sector in SECTOR_KEYWORDS:
            text = soup.get_text(separator=" ", strip=True).lower()
            if not any(kw in text for kw in SECTOR_KEYWORDS[self.sector]):
                log.debug(f"Descartado {url} por no coincidir con sector {self.sector}")
                return
                
        name = self._extract_name(soup, url)
        if not name or _is_blocked_name(name):
            return

        # Usar pipeline completo (mailto, JSON-LD, Cloudflare, regex)
        emails = self._extract_emails_from_soup(soup)
        phones = self._extract_phones_from_soup(soup)

        # Si faltan datos, seguir a página de contacto
        if not emails or not phones:
            cu = self._find_contact_url(url, soup)
            if cu and cu not in self.sources_visited:
                self.sources_visited.add(cu)
                cs = self._get(cu, referer=url)
                if cs:
                    if not emails:
                        emails = self._extract_emails_from_soup(cs)
                    if not phones:
                        phones = self._extract_phones_from_soup(cs)

        company = Company(
            empresa=name, sector=self.sector, web=url,
            email=emails[0] if emails else None,
            telefono=phones[0] if phones else None,
            fuente=source,
        )
        key = company.key()
        if key and key not in self.companies:
            self.companies[key] = company
            self._emit(f"✓ {name}", type="company", company=asdict(company), stats=self.stats)

    def _get_company_links_from_directory(self, url: str) -> list[str]:
        soup = self._get(url)
        if not soup:
            return []
        source_domain = urlparse(url).netloc.lower()
        seen_domains: set[str] = set()
        links = []
        bad_paths = ["/login","/logout","/signup","/register","/privacy",
                     "/terms","/cookies","/help","?logout=","shareArticle","/share?"]
        for a in soup.find_all("a", href=True):
            href = a.get("href","").strip()
            if not href or href.startswith("#") or href.startswith("mailto:"):
                continue
            full = urljoin(url, href)
            parsed = urlparse(full)
            if parsed.scheme not in ("http","https") or not parsed.netloc:
                continue
            d = parsed.netloc.lower()
            if d == source_domain or _is_blocked_domain(full):
                continue
            path_q = parsed.path + "?" + (parsed.query or "")
            if any(bp in path_q for bp in bad_paths):
                continue
            if d not in seen_domains:
                seen_domains.add(d)
                links.append(full)
            if len(links) >= 20:
                break
        return links

    def run(self):
        if self._stopped():
            return []
        queries = SECTOR_QUERIES.get(self.sector, SECTOR_QUERIES["Otros"])

        self._emit("🔍 Buscando directorios en España…", type="status", phase="searching")
        directory_urls: list[str] = []
        for query in queries:
            if self._stopped():
                break
            for url in self._search_urls(query)[:6]:
                if url not in directory_urls:
                    directory_urls.append(url)
            self._emit(f"Directorios: {len(directory_urls)}", type="status",
                       phase="searching", stats=self.stats)

        self._emit("🏢 Extrayendo empresas…", type="status", phase="extracting")
        for dir_url in directory_urls:
            if self._stopped():
                break
            self._emit(f"📂 {urlparse(dir_url).netloc}", type="status",
                       phase="extracting", stats=self.stats)
            for link in self._get_company_links_from_directory(dir_url):
                if self._stopped():
                    break
                self._process_company_url(link, source=dir_url)

        if not self._stopped():
            self._emit("📞 Enriqueciendo contactos…", type="status", phase="enriching")
            for company in [c for c in self.companies.values()
                            if not c.email and not c.telefono and c.web][:8]:
                if self._stopped():
                    break
                soup = self._get(company.web)
                if soup:
                    cu = self._find_contact_url(company.web, soup)
                    if cu:
                        self._process_company_url(cu, source=company.web)

        msg = "⏹ Detenido." if self._stopped() else "✅ Completado."
        ev_type = "stopped" if self._stopped() else "done"
        self._emit(msg, type=ev_type, stats=self.stats)
        return list(self.companies.values())


def export_results(companies: list[Company], output_dir: str = "output"):
    import csv, os
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    os.makedirs(output_dir, exist_ok=True)
    rows = [asdict(c) for c in companies]
    fields = ["empresa","sector","web","email","telefono","fuente"]

    csv_path = os.path.join(output_dir, "resultados.csv")
    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)

    xlsx_path = os.path.join(output_dir, "resultados.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Empresas"
    hf = PatternFill("solid", fgColor="0F172A")
    hfont = Font(color="22C55E", bold=True, name="Calibri", size=11)
    alt = PatternFill("solid", fgColor="F8FAFC")
    bdr = Border(bottom=Side(style="thin", color="E2E8F0"))
    ctr = Alignment(horizontal="center", vertical="center")
    headers = ["Empresa","Sector","Web","Email","Teléfono","Fuente"]
    widths  = [30, 15, 42, 35, 18, 45]
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = hfont; cell.fill = hf; cell.alignment = ctr
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 25
    for ri, row in enumerate(rows, 2):
        bg = PatternFill("solid", fgColor="FFFFFF") if ri%2==0 else alt
        for ci, f in enumerate(fields, 1):
            val = row.get(f) or ""
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill = bg; cell.border = bdr
            cell.alignment = Alignment(vertical="center")
            if ci == 3 and val:
                cell.hyperlink = val
                cell.font = Font(color="2563EB", underline="single")
    ws.freeze_panes = "A2"
    wb.save(xlsx_path)
    return csv_path, xlsx_path


DEMO_DATA = {
    "Tecnología": [
        Company("TechSolutions Madrid","Tecnología","https://techsolutions.es","info@techsolutions.es","+34912345678","https://directorio-tech.es"),
        Company("SoftDev Barcelona","Tecnología","https://softdev.barcelona","contacto@softdev.barcelona","+34934567890","https://directorio-tech.es"),
        Company("CloudInno Spain","Tecnología","https://cloudinno.es","hola@cloudinno.es",None,"https://startups-espana.com"),
        Company("DataPulse Analytics","Tecnología","https://datapulse.es","team@datapulse.es","+34912345678","https://empresas-tech.es"),
        Company("NexGen Software","Tecnología","https://nexgen.software",None,"+34931112233","https://directorio-software.es"),
    ],
    "Construcción": [
        Company("Constructora Ibérica","Construcción","https://constructoraiberica.es","info@constructoraiberica.es","+34914567890","https://directorio-construccion.es"),
        Company("Reformas García","Construcción","https://reformasgarcia.es","contacto@reformasgarcia.es","+34620345678","https://directorio-construccion.es"),
        Company("Edificaciones Modernas","Construcción","https://edificacionesmodernas.es",None,"+34935678901","https://construye.es"),
        Company("Grupo Constructor Norte","Construcción","https://grupocnorte.es","info@grupocnorte.es","+34944123456","https://construye.es"),
    ],
    "Marketing": [
        Company("Agencia CreativaX","Marketing","https://creativax.es","hola@creativax.es","+34918889900","https://agencias-marketing.es"),
        Company("Impulsa Digital","Marketing","https://impulsadigital.es","info@impulsadigital.es","+34937778899","https://directorio-marketing.es"),
        Company("Brand Studio BCN","Marketing","https://brandstudio.barcelona","studio@brandstudio.barcelona",None,"https://agencias-publicitarias.es"),
    ],
}

def get_demo_companies(sector: str) -> list[Company]:
    base = DEMO_DATA.get(sector, [])
    if not base:
        base = [Company(f"Empresa {sector} {i+1}", sector, f"https://empresa{i+1}.es",
                        f"info@empresa{i+1}.es", f"+3491{i}234567",
                        "https://directorio-ejemplo.es") for i in range(4)]
    return base
