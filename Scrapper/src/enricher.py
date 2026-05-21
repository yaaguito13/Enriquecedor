"""
enricher.py — Enriquecimiento especializado para ficheros RIPCI España
Cabeceras: Empresa, Contacto, Cargo, Teléfono x3, Correo x3,
           Sección, CCAA, Municipio, Provincia, Documento…
Los valores "PTE" se tratan como vacío pendiente de rellenar.
"""

import re, os, time, random, threading, logging, json as _json
from dataclasses import dataclass, asdict, field, fields as dc_fields
from typing import Optional
from urllib.parse import urlparse, urljoin, quote_plus, unquote

import requests
from bs4 import BeautifulSoup

log = logging.getLogger(__name__)

from scraper import (
    HEADERS, EMAIL_REGEX, PHONE_REGEX,
    _is_blocked_domain, _clean_email, _clean_phone,
    _make_headers, _decode_cloudflare_email,
    _extract_emails_from_obj, _extract_phones_from_obj,
    _extract_emails_from_text, _extract_phones_from_text,
    _normalize_url,
)

REQUEST_TIMEOUT = 12
DELAY = (1.2, 2.8)

# Valor "pendiente" en el Excel — tratar como vacío
PTE_VALUES = {"pte", "pendiente", "n/a", "#n/a", "-", "none", "null", "nan", ""}


def _is_empty(val) -> bool:
    """True si el valor es vacío, None o "PTE"."""
    if val is None:
        return True
    return str(val).strip().lower() in PTE_VALUES


def _existing_values(row, field_names: list[str]) -> set[str]:
    values = set()
    for field_name in field_names:
        raw = getattr(row, field_name)
        if _is_empty(raw):
            continue
        text = str(raw).strip()
        normalized = _clean_phone(text) if field_name.startswith("telefono") else _clean_email(text)
        values.add((normalized or text).lower())
    return values


def _fill_first_empty_slots(row, field_names: list[str], values: list[str]) -> bool:
    """Rellena huecos respetando datos existentes. Devuelve True si añade algo."""
    existing = _existing_values(row, field_names)
    changed = False
    for value in values:
        normalized = _clean_phone(value) if field_names[0].startswith("telefono") else _clean_email(value)
        value = normalized or value
        if not value or value.lower() in existing:
            continue
        for field_name in field_names:
            if _is_empty(getattr(row, field_name)):
                setattr(row, field_name, value)
                existing.add(value.lower())
                changed = True
                break
    return changed


# ──────────────────────────────────────────────────────────────
# Modelo de fila — campos específicos del Excel RIPCI
# ──────────────────────────────────────────────────────────────
@dataclass
class EnrichedRow:
    # ── Identificación ──────────────────────────────────────
    empresa:   str
    contacto:  Optional[str] = None  # Nombre del contacto
    cargo:     Optional[str] = None  # Cargo del contacto
    asesor:    Optional[str] = None  # Asesor asignado
    status:    Optional[str] = None  # Estado
    documento: Optional[str] = None  # NIF / CIF

    # ── Contacto (los campos a rellenar) ───────────────────
    telefono:  Optional[str] = None  # Teléfono principal
    telefono2: Optional[str] = None  # Teléfono 2
    email:     Optional[str] = None  # Email principal
    email2:    Optional[str] = None  # Email 2

    # ── Ubicación ───────────────────────────────────────────
    ccaa:      Optional[str] = None  # Comunidad Autónoma
    seccion:   Optional[str] = None  # Sección / Sector
    municipio: Optional[str] = None  # Municipio / Localidad
    provincia: Optional[str] = None  # Provincia
    zona:      Optional[str] = None  # Zona Geográfica

    # ── Web (no está en el original, la buscamos) ──────────
    web:       Optional[str] = None
    fuente:    Optional[str] = None  # Fuente del dato encontrado

    # ── Metadatos de enriquecimiento ───────────────────────
    email_nuevo:    bool = False
    telefono_nuevo: bool = False

    def needs_enrichment(self) -> bool:
        return _is_empty(self.email) or _is_empty(self.telefono)


# ──────────────────────────────────────────────────────────────
# Detección de columnas — mapa de alias para cabeceras RIPCI
# ──────────────────────────────────────────────────────────────
COLUMN_ALIASES = {
    "empresa":   ["empresa","company","nombre empresa","nombre de empresa","entidad","razon social","razón social","name"],
    "contacto":  ["contacto","nombre contacto","persona contacto","contact","contact name"],
    "cargo":     ["cargo","puesto","posición","position","role","título","title"],
    "asesor":    ["asesor","agente","responsable","gestor","agent"],
    "status":    ["status","estado","situación","state"],
    "documento": ["documento","nif","cif","nif/cif","vat","tax id","id fiscal","nif empresa"],
    "telefono":  ["teléfono","telefono","phone","tel","tlf","telf","fono","teléfono principal","tel principal","telephone"],
    "telefono2": ["teléfono 2","telefono 2","phone 2","tel 2","tlf 2","teléfono2","telefono2"],
    "email":     ["correo electrónico","correo electronico","email","e-mail","mail","correo","email principal","email contacto","correo principal"],
    "email2":    ["correo electrónico 2","correo electronico 2","email 2","e-mail 2","mail 2","correo 2","email2"],
    "ccaa":      ["ccaa","comunidad autónoma","comunidad autonoma","autonomous community","región","region"],
    "seccion":   ["sección","seccion","sector","industria","actividad","industry","category","categoría"],
    "municipio": ["municipio","localidad","municipio/localidad","ciudad","city","town","population"],
    "provincia": ["provincia","province"],
    "zona":      ["zona geográfica","zona geografica","zona","zone","area geográfica","area geografica"],
    "web":       ["web","url","website","pagina web","página web","sitio web","site","dominio","link"],
    "fuente":    ["fuente","source","origen"],
}


def detect_columns(headers: list[str]) -> dict[str, Optional[int]]:
    """Detecta qué columna corresponde a cada campo interno."""
    mapping: dict[str, Optional[int]] = {k: None for k in COLUMN_ALIASES}
    lower_h = [h.lower().strip() for h in headers]
    for field_name, aliases in COLUMN_ALIASES.items():
        # Primero intenta coincidencia exacta, luego parcial
        for i, h in enumerate(lower_h):
            if any(h == alias for alias in aliases):
                mapping[field_name] = i
                break
        if mapping[field_name] is None:
            for i, h in enumerate(lower_h):
                if any(alias in h for alias in aliases):
                    mapping[field_name] = i
                    break
    return mapping


# ──────────────────────────────────────────────────────────────
# Lectura de ficheros
# ──────────────────────────────────────────────────────────────
def read_file(path: str) -> tuple[list[str], list[list]]:
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xls", ".xlsm"):
        return _read_excel(path)
    elif ext in (".csv", ".tsv"):
        return _read_csv(path)
    raise ValueError(f"Formato no soportado: {ext}")


def _read_excel(path: str) -> tuple[list[str], list[list]]:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_raw = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows_raw:
        raise ValueError("El archivo Excel está vacío")
    headers = [str(c).strip() if c is not None else f"Col{i}"
               for i, c in enumerate(rows_raw[0])]
    rows = []
    for raw in rows_raw[1:]:
        row = [c for c in raw]  # Preservar tipos (int para teléfonos)
        if any(c is not None and str(c).strip() for c in row):
            while len(row) < len(headers):
                row.append(None)
            rows.append(row)
    return headers, rows


def _read_csv(path: str) -> tuple[list[str], list[list]]:
    import csv
    for enc in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
        try:
            with open(path, encoding=enc, newline="") as f:
                sample = f.read(4096)
            encoding = enc
            break
        except UnicodeDecodeError:
            continue
    else:
        encoding = "latin-1"
    with open(path, encoding=encoding, newline="") as f:
        sample = f.read(2048); f.seek(0)
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        reader = csv.reader(f, dialect)
        all_rows = list(reader)
    if not all_rows:
        raise ValueError("El CSV está vacío")
    headers = [h.strip() for h in all_rows[0]]
    rows = []
    for raw in all_rows[1:]:
        row = [c.strip() for c in raw]
        if any(row):
            while len(row) < len(headers):
                row.append("")
            rows.append(row)
    return headers, rows


def _get_val(raw: list, col: dict, field_name: str) -> Optional[str]:
    """Extrae y limpia un valor de una fila. Devuelve None si es PTE/vacío."""
    idx = col.get(field_name)
    if idx is None or idx >= len(raw):
        return None
    v = raw[idx]
    if v is None:
        return None
    s = str(v).strip()
    return None if s.lower() in PTE_VALUES else s


# ──────────────────────────────────────────────────────────────
# Agente de enriquecimiento
# ──────────────────────────────────────────────────────────────
class EnricherAgent:
    CONTACT_PATHS = [
        "/contacto", "/contact", "/contactar", "/es/contacto",
        "/sobre-nosotros/contacto", "/about/contact",
        "/quienes-somos", "/quienes-somos/contacto", "/empresa",
        "/aviso-legal", "/legal", "/privacy-policy",
        "/atencion-al-cliente", "/atencion-cliente",
        "/ayuda/contacto", "/help/contact",
        "/prensa", "/sala-de-prensa", "/delegaciones", "/oficinas",
    ]

    def __init__(self, status_cb=None):
        self.status_cb = status_cb or (lambda msg, **kw: None)
        self.session = requests.Session()
        self.session.headers.update(HEADERS)
        self._stop_event = threading.Event()
        self.results: list[EnrichedRow] = []

    def stop(self): self._stop_event.set()
    def _stopped(self) -> bool: return self._stop_event.is_set()
    def _emit(self, msg: str, **kw): self.status_cb(msg, **kw)

    @property
    def stats(self) -> dict:
        c = self.results
        return {
            "total":     len(c),
            "enriched":  sum(1 for r in c if r.email_nuevo or r.telefono_nuevo),
            "emails":    sum(1 for r in c if not _is_empty(r.email)),
            "telefonos": sum(1 for r in c if not _is_empty(r.telefono)),
        }

    # ── HTTP ──────────────────────────────────────────────────
    def _get(self, url: str) -> Optional[BeautifulSoup]:
        url = _normalize_url(url)
        if self._stopped() or _is_blocked_domain(url):
            return None
        try:
            self.session.headers.update(_make_headers(referer=url))
            time.sleep(random.uniform(*DELAY))
            resp = self.session.get(url, timeout=REQUEST_TIMEOUT, allow_redirects=True)
            if resp.status_code in (403, 429, 503):
                return None
            resp.raise_for_status()
            if _is_blocked_domain(resp.url) or len(resp.text.strip()) < 200:
                return None
            return BeautifulSoup(resp.text, "html.parser")
        except Exception as e:
            log.debug(f"GET {url}: {e}")
            return None

    # ── Extracción ────────────────────────────────────────────
    def _emails_from_text(self, text: str) -> list[str]:
        return _extract_emails_from_text(text)

    def _phones_from_text(self, text: str) -> list[str]:
        return _extract_phones_from_text(text)

    def _extract_from_soup(self, soup: BeautifulSoup) -> tuple[list[str], list[str]]:
        emails, phones = [], []
        # mailto / tel hrefs
        for a in soup.find_all("a", href=True):
            h = a["href"]
            if h.lower().startswith("mailto:"):
                c = _clean_email(h[7:].split("?")[0].strip())
                if c and c not in emails: emails.append(c)
            if h.lower().startswith("tel:"):
                c = _clean_phone(h[4:].strip())
                if c and c not in phones: phones.append(c)
        # JSON-LD
        for s in soup.find_all("script", type="application/ld+json"):
            try:
                data = _json.loads(s.string or "")
                for e in _extract_emails_from_obj(data):
                    c = _clean_email(e)
                    if c and c not in emails: emails.append(c)
                for p in _extract_phones_from_obj(data):
                    c = _clean_phone(p)
                    if c and c not in phones: phones.append(c)
            except Exception: pass
        # Cloudflare
        for el in soup.select("[data-cfemail],.__cf_email__"):
            dec = _decode_cloudflare_email(el.get("data-cfemail",""))
            if dec:
                c = _clean_email(dec)
                if c and c not in emails: emails.append(c)
        # Texto plano
        text = soup.get_text(" ", strip=True)
        for e in self._emails_from_text(text):
            if e not in emails: emails.append(e)
        for p in self._phones_from_text(text):
            if p not in phones: phones.append(p)
        # HTML crudo: captura contactos en atributos, JS inline y datos ocultos
        raw_html = str(soup)
        for e in self._emails_from_text(raw_html):
            if e not in emails: emails.append(e)
        for p in self._phones_from_text(raw_html):
            if p not in phones: phones.append(p)
        return emails, phones

    def _find_contact_url(self, base: str, soup: BeautifulSoup) -> Optional[str]:
        SCORES = {
            "contacto":10,"contact":10,"contactar":9,"contáctanos":9,
            "aviso legal":7,"sobre nosotros":6,"quiénes somos":6,"about":4,
        }
        cands = []
        for a in soup.find_all("a", href=True):
            href = a["href"].lower(); text = a.get_text(strip=True).lower()
            if any(x in href for x in ("#","javascript:","mailto:","tel:")):
                continue
            full = urljoin(base, a["href"])
            if _is_blocked_domain(full): continue
            score = sum(pts for kw,pts in SCORES.items() if kw in href or kw in text)
            if score > 0: cands.append((score, full))
        cands.sort(reverse=True)
        return cands[0][1] if cands else None

    def _try_contact_paths(self, base: str) -> tuple[list, list]:
        parsed = urlparse(base)
        root = f"{parsed.scheme}://{parsed.netloc}"
        emails, phones = [], []
        for path in self.CONTACT_PATHS:
            if self._stopped(): break
            soup = self._get(root + path)
            if not soup: continue
            e, p = self._extract_from_soup(soup)
            emails.extend(x for x in e if x not in emails)
            phones.extend(x for x in p if x not in phones)
            if emails and phones: break
        return emails, phones

    def _search_contact_info(self, name: str, municipio: str = "",
                              provincia: str = "", web: str = "") -> tuple[list, list]:
        """Busca contacto en DuckDuckGo usando nombre + ubicación."""
        from ddgs import DDGS
        emails, phones = [], []
        location = f"{municipio} {provincia}".strip()
        queries = [
            f'"{name}" teléfono contacto{" " + location if location else ""}',
            f'"{name}" email correo{" " + location if location else ""}',
        ]
        own_domain = urlparse(web).netloc if web else ""
        for query in queries:
            if self._stopped(): break
            try:
                with DDGS() as ddgs:
                    results = list(ddgs.text(query, region="es-es", max_results=4))
                    for res in results:
                        text = res.get("body", "") + " " + res.get("title", "")
                        for e in self._emails_from_text(text):
                            if e not in emails: emails.append(e)
                        for p in self._phones_from_text(text):
                            if p not in phones: phones.append(p)
                    if emails or phones: break

                    # Seguir primer resultado externo
                    for res in results:
                        href = res.get("href", "")
                        if href.startswith("http") and not _is_blocked_domain(href) and (not own_domain or own_domain not in href):
                            rs = self._get(href)
                            if rs:
                                e, p = self._extract_from_soup(rs)
                                emails.extend(x for x in e if x not in emails)
                                phones.extend(x for x in p if x not in phones)
                            break
            except Exception as e:
                log.debug(f"DDGS error en _search_contact_info: {e}")
        return emails, phones

    def _find_web(self, name: str, municipio: str = "", seccion: str = "") -> Optional[str]:
        """Busca la web oficial de la empresa en DuckDuckGo."""
        from ddgs import DDGS
        extra = " ".join(filter(None, [municipio, seccion]))
        q = f'"{name}" {extra} web oficial site:.es OR site:.com'
        try:
            with DDGS() as ddgs:
                for res in ddgs.text(q, region="es-es", max_results=3):
                    href = res.get("href", "")
                    if href.startswith("http") and not _is_blocked_domain(href):
                        return href
        except Exception:
            pass
        return None

    # ── Directorios públicos que tienen datos por CIF ──────────────
    # Se visitan directamente — bypass de la lista negra del scraper
    # porque aquí los usamos como FUENTE de datos, no como resultado
    _CIF_SOURCES = [
        "https://www.infocif.es/ficha-empresa/{cif}",
        "https://www.axesor.es/buscar?q={cif}",
        "https://einforma.com/informes/empresas/gratis/{cif}",
        "https://www.infoempresa.com/empresa/es/{cif}",
    ]

    def _get_cif_page(self, url: str) -> Optional[BeautifulSoup]:
        """GET sin filtro de dominio bloqueado — solo para directorios CIF."""
        if self._stopped():
            return None
        try:
            self.session.headers.update(_make_headers(referer=url))
            time.sleep(random.uniform(*DELAY))
            resp = self.session.get(url, timeout=REQUEST_TIMEOUT, allow_redirects=True)
            if resp.status_code in (403, 429, 503):
                return None
            resp.raise_for_status()
            if len(resp.text.strip()) < 200:
                return None
            return BeautifulSoup(resp.text, "html.parser")
        except Exception as e:
            log.debug(f"CIF GET {url}: {e}")
            return None

    def _search_by_cif(self, cif: str, empresa: str = "") -> tuple[list, list]:
        """
        Estrategia principal para el Excel RIPCI.
        Busca email y teléfono usando el CIF/NIF como identificador único.

        Flujo:
        1. DuckDuckGo snippet search: "{CIF}" teléfono contacto
        2. Directorios públicos directos: infocif, axesor, einforma, infoempresa
        3. DuckDuckGo + nombre empresa como refuerzo
        """
        if not cif or len(cif) < 7:
            return [], []

        cif = cif.strip().upper()
        emails, phones = [], []

        # ── 1. Búsqueda DuckDuckGo por CIF (snippets) ──────────────
        from ddgs import DDGS
        for query in [
            f'"{cif}" teléfono contacto empresa',
            f'"{cif}" correo email empresa',
        ]:
            if self._stopped(): break
            try:
                with DDGS() as ddgs:
                    results = list(ddgs.text(query, region="es-es", max_results=4))
                    for res in results:
                        text = res.get("body", "") + " " + res.get("title", "")
                        for e in self._emails_from_text(text):
                            if e not in emails: emails.append(e)
                        for p in self._phones_from_text(text):
                            if p not in phones: phones.append(p)
                            
                    if emails and phones: return emails, phones

                    # Seguir los primeros 2 resultados de búsqueda
                    links_followed = 0
                    for res in results:
                        if links_followed >= 2 or (emails and phones): break
                        href = res.get("href", "")
                        if not href.startswith("http"): continue
                        rs = self._get_cif_page(href)
                        if rs:
                            ce, cp = self._extract_from_soup(rs)
                            emails.extend(x for x in ce if x not in emails)
                            phones.extend(x for x in cp if x not in phones)
                            links_followed += 1

                    if emails and phones: return emails, phones
            except Exception as e:
                log.debug(f"DDGS error en _search_by_cif (1): {e}")

        # ── 2. Directorios públicos directos por CIF ──────────────
        for url_tpl in self._CIF_SOURCES:
            if self._stopped() or (emails and phones): break
            url = url_tpl.format(cif=cif)
            soup = self._get_cif_page(url)
            if not soup: continue
            ce, cp = self._extract_from_soup(soup)
            emails.extend(x for x in ce if x not in emails)
            phones.extend(x for x in cp if x not in phones)
            if emails and phones: break

        # ── 3. Búsqueda combinada CIF + nombre (refuerzo) ──────────
        if (not emails or not phones) and empresa:
            q = f'"{cif}" "{empresa}" contacto teléfono email'
            try:
                with DDGS() as ddgs:
                    for res in ddgs.text(q, region="es-es", max_results=3):
                        text = res.get("body", "") + " " + res.get("title", "")
                        for e in self._emails_from_text(text):
                            if e not in emails: emails.append(e)
                        for p in self._phones_from_text(text):
                            if p not in phones: phones.append(p)
            except Exception as e:
                log.debug(f"DDGS error en _search_by_cif (3): {e}")

        return emails, phones

    def _scrape_web(self, url: str, name: str = "",
                    municipio: str = "", provincia: str = "") -> tuple[list, list]:
        """3 estrategias: scraping directo → rutas comunes → búsqueda DDG."""
        emails, phones = [], []
        url = _normalize_url(url)
        if url and not _is_blocked_domain(url):
            soup = self._get(url)
            if soup:
                emails, phones = self._extract_from_soup(soup)
                if not emails or not phones:
                    cu = self._find_contact_url(url, soup)
                    if cu:
                        cs = self._get(cu)
                        if cs:
                            ce, cp = self._extract_from_soup(cs)
                            emails.extend(x for x in ce if x not in emails)
                            phones.extend(x for x in cp if x not in phones)
            if not emails or not phones:
                ce, cp = self._try_contact_paths(url)
                emails.extend(x for x in ce if x not in emails)
                phones.extend(x for x in cp if x not in phones)
        if (not emails or not phones) and name:
            ce, cp = self._search_contact_info(name, municipio, provincia, web=url)
            emails.extend(x for x in ce if x not in emails)
            phones.extend(x for x in cp if x not in phones)
        return emails, phones

    # ── Flujo principal ───────────────────────────────────────
    def run(self, file_path: str) -> list[EnrichedRow]:
        self._emit("📂 Leyendo fichero…", type="status", phase="reading")
        try:
            headers, rows = read_file(file_path)
        except Exception as e:
            self._emit(f"❌ {e}", type="error"); return []
        if not rows:
            self._emit("❌ Fichero vacío", type="error"); return []

        col = detect_columns(headers)
        self._emit(
            f"✓ {len(rows)} filas | columnas: "
            + ", ".join(f"{k}" for k,v in col.items() if v is not None),
            type="status", phase="reading",
        )
        if col["empresa"] is None:
            self._emit("❌ No se detectó columna 'Empresa'", type="error"); return []

        # ── Construir EnrichedRow por fila ────────────────────
        enriched: list[EnrichedRow] = []
        for raw in rows:
            g = lambda f: _get_val(raw, col, f)
            enriched.append(EnrichedRow(
                empresa   = g("empresa") or f"Fila {len(enriched)+1}",
                contacto  = g("contacto"),
                cargo     = g("cargo"),
                asesor    = g("asesor"),
                status    = g("status"),
                documento = g("documento"),
                telefono  = g("telefono"),
                telefono2 = g("telefono2"),
                email     = g("email"),
                email2    = g("email2"),
                ccaa      = g("ccaa"),
                seccion   = g("seccion"),
                municipio = g("municipio"),
                provincia = g("provincia"),
                zona      = g("zona"),
                web       = g("web"),
            ))

        total  = len(enriched)
        need   = sum(1 for r in enriched if r.needs_enrichment())

        # Comprobar modo AI
        try:
            from ai_enricher import check_api_key, enrich_with_ai
            api_ok, _ = check_api_key()
        except Exception:
            api_ok = False

        modo = "🤖 IA + Web Search" if api_ok else "🔍 Scraping en cascada"
        self._emit(
            f"📊 {total} filas | {need} necesitan enriquecimiento | Modo: {modo}",
            type="summary",
            summary={"total": total, "need": need},
            stats={"total": total, "enriched": 0, "emails": 0, "telefonos": 0},
        )

        self._emit("🔍 Iniciando enriquecimiento…", type="status", phase="enriching")

        for i, row in enumerate(enriched):
            if self._stopped(): break

            if not row.needs_enrichment():
                self.results.append(row)
                self._emit(f"✓ {row.empresa} — completo",
                           type="row", row=asdict(row), index=i, stats=self.stats)
                continue

            email_fields = ["email", "email2"]
            phone_fields = ["telefono", "telefono2"]
            need_email = _is_empty(row.email)
            need_phone = _is_empty(row.telefono)
            self._emit(
                f"🔎 [{i+1}/{total}] {row.empresa} "
                f"({'email+tel' if need_email and need_phone else 'email' if need_email else 'tel'})",
                type="status", phase="enriching", stats=self.stats,
            )

            # Buscar web si no la tiene
            if not row.web:
                row.web = self._find_web(
                    row.empresa,
                    municipio=row.municipio or "",
                    seccion=row.seccion or "",
                )
                if row.web:
                    self._emit(f"   ↳ Web: {row.web}", type="status", phase="enriching")

            emails_found: list[str] = []
            phones_found: list[str] = []
            fuente = None

            # ── Estrategia 0: Búsqueda por CIF (más precisa) ─────
            # El CIF es único — da resultados mucho más exactos que buscar por nombre
            if row.documento and (need_email or need_phone):
                self._emit(f"   ↳ 🔑 CIF: {row.documento}…",
                           type="status", phase="enriching")
                ce, cp = self._search_by_cif(row.documento, empresa=row.empresa)
                if need_email and ce:
                    emails_found.extend(x for x in ce if x not in emails_found)
                    fuente = "CIF-Directorio"
                if need_phone and cp:
                    phones_found.extend(x for x in cp if x not in phones_found)
                    fuente = fuente or "CIF-Directorio"
                if emails_found or phones_found:
                    self._emit(
                        f"   ↳ ✓ CIF encontró: "
                        + (f"{len(emails_found)} email(s) " if emails_found else "")
                        + (f"{len(phones_found)} teléfono(s)" if phones_found else ""),
                        type="status", phase="enriching",
                    )

            # ── Modo AI (si falta algo tras búsqueda CIF) ────────
            if api_ok and ((need_email and len(emails_found) < 2) or (need_phone and len(phones_found) < 2)):
                try:
                    self._emit("   ↳ 🤖 IA buscando lo que falta…",
                               type="status", phase="enriching")
                    res = enrich_with_ai(
                        empresa=row.empresa,
                        web=row.web or "",
                        sector=row.seccion or "",
                        need_email=(need_email and len(emails_found) < 2),
                        need_phone=(need_phone and len(phones_found) < 2),
                    )
                    if need_email and res.get("email") and res["email"] not in emails_found:
                        emails_found.append(res["email"]); fuente = fuente or "AI+WebSearch"
                    if need_phone and res.get("telefono") and res["telefono"] not in phones_found:
                        phones_found.append(res["telefono"]); fuente = fuente or "AI+WebSearch"
                except Exception as e:
                    self._emit(f"   ↳ ⚠️ IA falló: {e}", type="status", phase="enriching")

            # ── Scraping cascada (último recurso) ─────────────────
            if (need_email and len(emails_found) < 2) or (need_phone and len(phones_found) < 2):
                se, sp = self._scrape_web(
                    row.web or "",
                    name=row.empresa,
                    municipio=row.municipio or "",
                    provincia=row.provincia or "",
                )
                if need_email and se:
                    emails_found.extend(x for x in se if x not in emails_found)
                    fuente = fuente or "Scraping"
                if need_phone and sp:
                    phones_found.extend(x for x in sp if x not in phones_found)
                    fuente = fuente or "Scraping"

            # ── Guardar ───────────────────────────────────────
            if need_email and _fill_first_empty_slots(row, email_fields, emails_found[:2]):
                row.email_nuevo = True
            if need_phone and _fill_first_empty_slots(row, phone_fields, phones_found[:2]):
                row.telefono_nuevo = True
            if fuente and not row.fuente:
                row.fuente = fuente

            nuevo = []
            if row.email_nuevo:    nuevo.append("email")
            if row.telefono_nuevo: nuevo.append("teléfono")
            icon = "✅" if nuevo else ("⚠️" if row.needs_enrichment() else "✓")
            msg  = f"{icon} {row.empresa}"
            if nuevo: msg += f" — nuevo: {', '.join(nuevo)}"

            self.results.append(row)
            self._emit(msg, type="row", row=asdict(row), index=i,
                       new_fields=nuevo, stats=self.stats)
            if api_ok and not self._stopped(): time.sleep(0.8)

        status = "⏹ Detenido." if self._stopped() else \
                 f"✅ Completado — {self.stats['enriched']} enriquecidas"
        ev = "stopped" if self._stopped() else "done"
        self._emit(status, type=ev, stats=self.stats)
        return self.results


# ──────────────────────────────────────────────────────────────
# Exportación adaptada al formato RIPCI
# ──────────────────────────────────────────────────────────────
def export_enriched(rows: list[EnrichedRow], output_dir: str = "output") -> tuple[str, str]:
    import csv
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    os.makedirs(output_dir, exist_ok=True)

    # Columnas de salida en el orden RIPCI original + web + fuente al final
    COLUMNS = [
        ("Empresa",              "empresa"),
        ("Contacto",             "contacto"),
        ("Cargo",                "cargo"),
        ("Asesor",               "asesor"),
        ("Status",               "status"),
        ("Documento (NIF/CIF)",  "documento"),
        ("Teléfono",             "telefono"),
        ("Teléfono 2",           "telefono2"),
        ("Correo electrónico",   "email"),
        ("Correo electrónico 2", "email2"),
        ("CCAA",                 "ccaa"),
        ("Sección",              "seccion"),
        ("Municipio/Localidad",  "municipio"),
        ("Provincia",            "provincia"),
        ("Zona Geográfica",      "zona"),
        ("Web",                  "web"),
        ("Fuente dato",          "fuente"),
    ]

    raw_rows = [asdict(r) for r in rows]
    fields_out = [f for _, f in COLUMNS]

    # ── CSV ──────────────────────────────────────────────────
    csv_path = os.path.join(output_dir, "enriquecido.csv")
    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=fields_out + ["email_nuevo","telefono_nuevo"],
                           extrasaction="ignore")
        w.writeheader(); w.writerows(raw_rows)

    # ── Excel ────────────────────────────────────────────────
    xlsx_path = os.path.join(output_dir, "enriquecido.xlsx")
    wb = Workbook(); ws = wb.active; ws.title = "RIPCI Enriquecido"

    HDR_FILL  = PatternFill("solid", fgColor="0F172A")
    HDR_FONT  = Font(color="22C55E", bold=True, name="Calibri", size=10)
    NEW_FILL  = PatternFill("solid", fgColor="DCFCE7")   # verde = nuevo
    NEW_FONT  = Font(color="166534", bold=True, name="Calibri", size=10)
    ALT_FILL  = PatternFill("solid", fgColor="F8FAFC")
    BDR       = Border(bottom=Side(style="thin", color="E2E8F0"))
    CTR       = Alignment(horizontal="center", vertical="center", wrap_text=False)

    widths = [35,20,18,12,10,16,16,35,35,18,20,22,18,18,35,22]

    for ci, ((hdr, _), w) in enumerate(zip(COLUMNS, widths), 1):
        cell = ws.cell(row=1, column=ci, value=hdr)
        cell.font = HDR_FONT; cell.fill = HDR_FILL
        cell.alignment = CTR
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 22

    for ri, raw in enumerate(raw_rows, 2):
        bg = PatternFill("solid", fgColor="FFFFFF") if ri%2==0 else ALT_FILL
        for ci, (_, field_name) in enumerate(COLUMNS, 1):
            val  = raw.get(field_name) or ""
            is_new = (field_name == "email"    and raw.get("email_nuevo")) or \
                     (field_name == "telefono" and raw.get("telefono_nuevo"))
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill   = NEW_FILL if is_new else bg
            cell.font   = NEW_FONT if is_new else Font(name="Calibri", size=10)
            cell.border = BDR
            cell.alignment = Alignment(vertical="center")
            if field_name == "web" and val:
                cell.hyperlink = val
                cell.font = Font(color="2563EB", underline="single",
                                 name="Calibri", size=10)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    # Leyenda
    leg_row = len(raw_rows) + 3
    c = ws.cell(row=leg_row, column=1, value="🟢 Celdas en verde = datos añadidos por EmpresaScout")
    c.font = Font(color="16A34A", italic=True, size=9, name="Calibri")

    wb.save(xlsx_path)
    return csv_path, xlsx_path
