from __future__ import annotations

from collections.abc import Iterable
from dataclasses import dataclass

from openpyxl.worksheet.worksheet import Worksheet

from company_enricher.core.models import ColumnMapping
from company_enricher.core.text import normalize_text
from company_enricher.core.validators import EMAIL_RE, is_probably_website, normalize_spanish_phone


FIELD_KEYWORDS = {
    "company": {
        "empresa", "razon social", "razon", "nombre fiscal", "nombre comercial",
        "cliente", "compania", "sociedad", "denominacion", "entidad", "proveedor",
    },
    "phone": {"telefono", "tel", "fijo", "centralita", "contacto telefono", "phone"},
    "email": {"email", "e mail", "correo", "mail", "correo electronico"},
    "website": {"web", "website", "pagina web", "url", "site", "dominio"},
    "address": {"direccion", "domicilio", "calle", "address"},
    "city": {"ciudad", "localidad", "municipio", "poblacion"},
    "province": {"provincia", "region", "comunidad"},
}


@dataclass
class HeaderCandidate:
    row_index: int
    headers: list[str]
    score: float


def detect_header_row(ws: Worksheet, max_rows: int = 12) -> HeaderCandidate:
    best = HeaderCandidate(row_index=1, headers=[], score=-1)
    for row_index in range(1, min(ws.max_row, max_rows) + 1):
        values = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[row_index]]
        non_empty = [value for value in values if value]
        keyword_hits = sum(_header_keyword_score(value) for value in non_empty)
        uniqueness = len(set(non_empty)) / max(len(non_empty), 1)
        score = keyword_hits + len(non_empty) * 0.05 + uniqueness
        if score > best.score:
            best = HeaderCandidate(row_index=row_index, headers=values, score=score)
    return best


def detect_columns(ws: Worksheet, header_row: int | None = None) -> tuple[int, ColumnMapping]:
    candidate = detect_header_row(ws) if header_row is None else HeaderCandidate(header_row, [], 0)
    header_index = candidate.row_index
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[header_index]]
    mapping = ColumnMapping()

    for field in FIELD_KEYWORDS:
        best_col, best_score = _best_header_match(headers, field)
        value_score = _best_value_match(ws, header_index, field)
        if value_score[1] > best_score:
            best_col, best_score = value_score
        if best_col and best_score >= 0.35:
            setattr(mapping, field, best_col)
            mapping.confidence[field] = round(best_score, 3)

    if not mapping.company:
        mapping.company = _fallback_company_column(ws, header_index, headers)
        if mapping.company:
            mapping.confidence["company"] = 0.3

    return header_index, mapping


def _best_header_match(headers: list[str], field: str) -> tuple[str | None, float]:
    best_name: str | None = None
    best_score = 0.0
    for header in headers:
        if not header:
            continue
        normalized = normalize_text(header)
        score = max((_keyword_similarity(normalized, kw) for kw in FIELD_KEYWORDS[field]), default=0.0)
        if score > best_score:
            best_name = header
            best_score = score
    return best_name, best_score


def _best_value_match(ws: Worksheet, header_row: int, field: str) -> tuple[str | None, float]:
    best_header: str | None = None
    best_score = 0.0
    for column_index, header_cell in enumerate(ws[header_row], start=1):
        header = str(header_cell.value).strip() if header_cell.value else f"Columna {column_index}"
        samples = _sample_column(ws, column_index, header_row + 1)
        if not samples:
            continue
        score = _value_score(samples, field)
        if score > best_score:
            best_header = header
            best_score = score
    return best_header, best_score


def _sample_column(ws: Worksheet, column_index: int, start_row: int, limit: int = 30) -> list[str]:
    values: list[str] = []
    for row in range(start_row, min(ws.max_row, start_row + limit - 1) + 1):
        value = ws.cell(row=row, column=column_index).value
        if value is not None and str(value).strip():
            values.append(str(value).strip())
    return values


def _value_score(samples: Iterable[str], field: str) -> float:
    values = list(samples)
    if not values:
        return 0.0
    if field == "email":
        return sum(1 for value in values if EMAIL_RE.search(value)) / len(values)
    if field == "phone":
        return sum(1 for value in values if normalize_spanish_phone(value)) / len(values)
    if field == "website":
        return sum(1 for value in values if is_probably_website(value)) / len(values)
    if field == "company":
        texty = sum(1 for value in values if any(char.isalpha() for char in value) and len(value) > 2)
        too_long = sum(1 for value in values if len(value) > 90)
        return max(0.0, (texty - too_long) / len(values) * 0.45)
    return 0.0


def _header_keyword_score(value: str) -> float:
    normalized = normalize_text(value)
    return max(
        (_keyword_similarity(normalized, keyword) for keywords in FIELD_KEYWORDS.values() for keyword in keywords),
        default=0.0,
    )


def _keyword_similarity(header: str, keyword: str) -> float:
    keyword = normalize_text(keyword)
    if not header or not keyword:
        return 0.0
    if header == keyword:
        return 1.0
    if keyword in header or header in keyword:
        return 0.78
    header_tokens = set(header.split())
    keyword_tokens = set(keyword.split())
    overlap = len(header_tokens & keyword_tokens)
    return overlap / max(len(keyword_tokens), 1) * 0.65


def _fallback_company_column(ws: Worksheet, header_row: int, headers: list[str]) -> str | None:
    best_header: str | None = None
    best_score = 0.0
    for column_index, header in enumerate(headers, start=1):
        samples = _sample_column(ws, column_index, header_row + 1)
        score = _value_score(samples, "company")
        if score > best_score:
            best_header = header or f"Columna {column_index}"
            best_score = score
    return best_header
