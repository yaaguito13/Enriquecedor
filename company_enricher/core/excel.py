from __future__ import annotations

import csv
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from company_enricher.core.column_detector import detect_columns
from company_enricher.core.models import ColumnMapping, CompanyRecord, EnrichmentResult, ExportArtifacts


ENRICHED_HEADERS = [
    "telefono_encontrado",
    "email_encontrado",
    "web_encontrada",
    "fuente_contacto",
    "confianza_contacto",
    "estado_enriquecimiento",
    "error_enriquecimiento",
]


def load_records(path: Path) -> tuple[int, ColumnMapping, list[CompanyRecord]]:
    wb = load_workbook(path)
    ws = wb.active
    header_row, mapping = detect_columns(ws)
    if not mapping.company:
        raise ValueError("No se pudo detectar la columna de empresa.")

    header_to_index = _headers_by_name(ws, header_row)
    company_idx = header_to_index[mapping.company]
    records: list[CompanyRecord] = []
    for row in range(header_row + 1, ws.max_row + 1):
        company_name = ws.cell(row=row, column=company_idx).value
        if company_name is None or not str(company_name).strip():
            continue
        raw = {
            str(ws.cell(row=header_row, column=col).value or f"Columna {col}"): ws.cell(row=row, column=col).value
            for col in range(1, ws.max_column + 1)
        }
        records.append(
            CompanyRecord(
                row_number=row,
                company_name=str(company_name).strip(),
                existing_phone=_cell_by_mapping(ws, row, header_to_index, mapping.phone),
                existing_email=_cell_by_mapping(ws, row, header_to_index, mapping.email),
                website=_cell_by_mapping(ws, row, header_to_index, mapping.website),
                address=_cell_by_mapping(ws, row, header_to_index, mapping.address),
                city=_cell_by_mapping(ws, row, header_to_index, mapping.city),
                province=_cell_by_mapping(ws, row, header_to_index, mapping.province),
                raw=raw,
            )
        )
    return header_row, mapping, records


def export_results(
    input_path: Path,
    output_dir: Path,
    job_id: str,
    header_row: int,
    results: list[EnrichmentResult],
    log_file: Path,
) -> ExportArtifacts:
    output_dir.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(input_path)
    ws = wb.active

    header_positions = _ensure_enriched_headers(ws, header_row)
    for result in results:
        row = result.record.row_number
        ws.cell(row=row, column=header_positions["telefono_encontrado"]).value = result.phone
        ws.cell(row=row, column=header_positions["email_encontrado"]).value = result.email
        ws.cell(row=row, column=header_positions["web_encontrada"]).value = result.website
        ws.cell(row=row, column=header_positions["fuente_contacto"]).value = result.source_url
        ws.cell(row=row, column=header_positions["confianza_contacto"]).value = round(result.confidence, 2)
        ws.cell(row=row, column=header_positions["estado_enriquecimiento"]).value = result.status
        ws.cell(row=row, column=header_positions["error_enriquecimiento"]).value = result.error

    enriched_path = output_dir / f"{job_id}_enriquecido.xlsx"
    errors_path = output_dir / f"{job_id}_errores.csv"
    not_found_path = output_dir / f"{job_id}_no_encontradas.csv"
    wb.save(enriched_path)
    _write_errors(errors_path, results)
    _write_not_found(not_found_path, results)
    return ExportArtifacts(enriched_excel=enriched_path, errors_csv=errors_path, not_found_csv=not_found_path, log_file=log_file)


def result_to_preview(result: EnrichmentResult) -> dict[str, Any]:
    return {
        "empresa": result.record.company_name,
        "telefono": result.phone or "",
        "email": result.email or "",
        "web": result.website or "",
        "estado": result.status,
        "confianza": round(result.confidence * 100, 1),
        "fuente": result.source_url or "",
        "error": result.error or "",
    }


def _headers_by_name(ws: Worksheet, header_row: int) -> dict[str, int]:
    headers: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        name = str(ws.cell(row=header_row, column=col).value or f"Columna {col}").strip()
        headers.setdefault(name, col)
    return headers


def _cell_by_mapping(ws: Worksheet, row: int, header_to_index: dict[str, int], header: str | None) -> str | None:
    if not header or header not in header_to_index:
        return None
    value = ws.cell(row=row, column=header_to_index[header]).value
    return str(value).strip() if value is not None and str(value).strip() else None


def _ensure_enriched_headers(ws: Worksheet, header_row: int) -> dict[str, int]:
    existing = _headers_by_name(ws, header_row)
    positions: dict[str, int] = {}
    next_col = ws.max_column + 1
    for header in ENRICHED_HEADERS:
        if header in existing:
            positions[header] = existing[header]
            continue
        positions[header] = next_col
        source_col = max(1, next_col - 1)
        ws.cell(row=header_row, column=next_col).value = header
        ws.cell(row=header_row, column=next_col)._style = ws.cell(row=header_row, column=source_col)._style
        ws.column_dimensions[get_column_letter(next_col)].width = max(18, ws.column_dimensions[get_column_letter(source_col)].width or 12)
        next_col += 1
    return positions


def _write_errors(path: Path, results: list[EnrichmentResult]) -> None:
    rows = [result for result in results if result.error]
    _write_csv(path, rows)


def _write_not_found(path: Path, results: list[EnrichmentResult]) -> None:
    rows = [result for result in results if result.status == "not_found"]
    _write_csv(path, rows)


def _write_csv(path: Path, rows: list[EnrichmentResult]) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=["fila", "empresa", "estado", "error", "fuente"])
        writer.writeheader()
        for result in rows:
            writer.writerow(
                {
                    "fila": result.record.row_number,
                    "empresa": result.record.company_name,
                    "estado": result.status,
                    "error": result.error or "",
                    "fuente": result.source_url or "",
                }
            )
